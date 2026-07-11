import openpyxl
import csv
from datetime import datetime

# Load workbook
wb = openpyxl.load_workbook(r"c:\xampp\htdocs\console\Console\Accounts\api\interns_payments_march.xlsx")
sheet = wb.active

csv_file = r"c:\xampp\htdocs\console\Console\Accounts\api\interns_payments_march.csv"

with open(csv_file, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    # Write header
    writer.writerow(["BillToName", "Phone", "Email", "Description", "PaymentMode", "Date", "CourseFee", "PaidNow", "Type"])
    
    # Process rows (start from row 2)
    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(row=r, column=1).value
        phone = sheet.cell(row=r, column=2).value
        total_amt = sheet.cell(row=r, column=3).value
        paid_date = sheet.cell(row=r, column=4).value
        paid_amt = sheet.cell(row=r, column=5).value
        paid_mode = sheet.cell(row=r, column=6).value
        
        # Skip empty rows or header duplicate rows
        if not name or not phone or str(name).strip() == "Student Name":
            continue
            
        # Clean phone
        phone_str = str(phone).strip()
        if phone_str.endswith(".0"):
            phone_str = phone_str[:-2]
            
        # Format Date
        date_str = ""
        if isinstance(paid_date, datetime):
            date_str = paid_date.strftime("%Y-%m-%d")
        elif paid_date:
            date_str = str(paid_date).strip()
            
        # Email placeholder
        email = ""
        
        # Total / Paid amounts
        try:
            total_amt_val = float(total_amt) if total_amt else 5000.0
        except ValueError:
            total_amt_val = 5000.0
            
        try:
            paid_amt_val = float(paid_amt) if paid_amt else 0.0
        except ValueError:
            paid_amt_val = 0.0
        
        # Mode
        mode_str = str(paid_mode).strip() if paid_mode else "Online"
        
        writer.writerow([
            str(name).strip(),
            phone_str,
            email,
            "Internship + project charges",
            mode_str,
            date_str,
            total_amt_val,
            paid_amt_val,
            "non-gst"
        ])

print("Conversion complete!")
