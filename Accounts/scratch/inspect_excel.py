import openpyxl

wb = openpyxl.load_workbook(r"c:\xampp\htdocs\console\Console\Accounts\api\interns_payments_march.xlsx")
sheet = wb.active

print("Sheet Title:", sheet.title)
print("Max Rows:", sheet.max_row)
print("Max Cols:", sheet.max_column)

for r in range(1, min(25, sheet.max_row + 1)):
    row_vals = [cell.value for cell in sheet[r]]
    if any(row_vals):
        print(f"Row {r}: {row_vals}")
